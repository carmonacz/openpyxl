from openpyxl import Workbook

libro = Workbook()
hoja = libro.active
#Insertamos valores a las celdas
for elem in range(1,6):
    hoja.cell(row = elem, column = 1).value = elem

#Insertamos una nueva fila
hoja.insert_rows(2)
#Insertamos un nuevo valor a la fila nueva
hoja.cell(row=2, column=1).value = 758
#Comprobamos el valor de la fila 4
print(hoja.cell(row=4,column=1).value)
#Modificamos el valor de la fila 4
hoja.cell(row=4, column=1).value = 5261
print(hoja.cell(row=4,column=1).value)
libro.save('demoInsertR.xlsx')